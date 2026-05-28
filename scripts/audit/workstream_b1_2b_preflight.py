"""Workstream B1.2b pre-flight coverage check (Turn N+5, Reese 2026-05-27).

Sample 5 schools from the 43 missing 1A-5A canonical schools, probe
lhsaaonline.org for 2021-2024 historical game records, and verdict
whether full-backfill Option (a) is viable vs falling back to hybrid
Option (c).

Pass criterion (Reese 2026-05-27): ≥80% of sampled schools have ≥80%
game coverage in 2021-2024.

Method
------
For each sample school:
  - Determine which sports the school fields per the LHSAA Schools by
    Sport.xlsx (commit 8ae9050) participation matrix.
  - For each (school, sport, year) tuple in years 2021-2024:
      - Compute the lhsaaonline filter_value the school would appear under
        (classification for most sports; division for volleyball).
      - Fetch one sport×year×filter_value report from lhsaaonline.
      - Search the response body for the school's name + count games.
  - Coverage % = sport-years with games found / sport-years probed.

Fetches are cached per (sport, year, filter_value) so multiple schools
with the same class share fetches. REQUEST_DELAY = 0.4s sustains the
existing scraper's rate-limit discipline.

Pass criterion check at end: count schools with coverage ≥ 80%; if
that count is ≥ 80% of samples (= 4 of 5), Option (a) locks.
"""
from __future__ import annotations

import csv
import json
import re
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "packages" / "engine" / "src"))

import httpx
import openpyxl
from bs4 import BeautifulSoup


REQUEST_DELAY = 0.4  # match ingest_sports_historical.py

# LHSAA per-sport URL/filter configuration (from
# scripts/ingest_sports_historical.py — abbreviated to what the pre-flight
# needs: filter style + base + form + report).
SPORT_CFG = {
    "Volleyball": {
        "base_url": "https://www.lhsaaonline.org/pr/vbpr/admin/",
        "form_path": "SearchVolleyballSchedule.asp",
        "report_suffix": "?p=1",
        "year_field": "y", "filter_field": "d",
        "filter_style": "division",   # I, II, III, IV, V
    },
    "Boys Basketball": {
        "base_url": "https://www.lhsaaonline.org/pr/bbpr/admin/",
        "form_path": "SearchBoysBasketballSchedule.asp",
        "report_suffix": "?p=1&bb=1",
        "year_field": "yr", "filter_field": "d",
        "filter_style": "class",      # 5A, 4A, 3A, 2A, 1A
    },
    "Girls Basketball": {
        "base_url": "https://www.lhsaaonline.org/pr/bbpr/admin/",
        "form_path": "SearchGirlsBasketballSchedule.asp",
        "report_suffix": "?p=1&bb=2",
        "year_field": "yr", "filter_field": "d",
        "filter_style": "class",
    },
    "Baseball": {
        "base_url": "https://www.lhsaaonline.org/pr/bpr/admin/",
        "form_path": "SearchBaseballSchedule.asp",
        "report_suffix": "?p=1&bb=1",
        "year_field": "y", "filter_field": "d",
        "filter_style": "class",
    },
    "Softball": {
        "base_url": "https://www.lhsaaonline.org/pr/sbpr/admin/",
        "form_path": "SearchSoftballSchedule.asp",
        "report_suffix": "?p=1&bb=2",
        "year_field": "y", "filter_field": "d",
        "filter_style": "class",
    },
    "Boys Soccer": {
        "base_url": "https://www.lhsaaonline.org/pr/sopr/admin/",
        "form_path": "SearchboyssoccerSchedule.asp",
        "report_suffix": "?p=1&so=1",
        "year_field": "yr", "filter_field": "d",
        "filter_style": "class",
    },
    "Girls Soccer": {
        "base_url": "https://www.lhsaaonline.org/pr/sopr/admin/",
        "form_path": "SearchgirlssoccerSchedule.asp",
        "report_suffix": "?p=1&so=2",
        "year_field": "yr", "filter_field": "d",
        "filter_style": "class",
    },
    # Football not in B1.2b scope (already at 100% coverage) but kept here
    # in case a sampled school fields football for completeness.
    "Football": {
        "base_url": "https://www.lhsaaonline.org/pr/fbpr/admin/",
        "form_path": "SearchFootballSchedule.asp",
        "report_suffix": "?p=1",
        "year_field": "yr", "filter_field": "d",
        "filter_style": "class",
    },
}

# CLASS_TO_DIV mapping for volleyball (per scraper's deprecated-but-still-used helper)
CLASS_TO_DIV = {"5A": "I", "4A": "II", "3A": "III", "2A": "IV", "1A": "V", "B": "V", "C": "V"}

YEARS = [2021, 2022, 2023, 2024]


# The 5 sample schools picked for diversity (1 per class 1A-5A, mix of
# cities, mix of expected sport-mixes).
SAMPLE_SCHOOLS = [
    ("Cohen College Prep", "New Orleans", "1A"),       # 1A NOLA charter
    ("Pickering", "Leesville", "2A"),                  # 2A rural
    ("Morris Jeff", "New Orleans", "3A"),              # 3A NOLA charter
    ("Caddo Magnet", "Shreveport", "4A"),              # 4A Shreveport magnet
    ("Mt. Carmel", "New Orleans", "5A"),               # 5A NOLA Catholic
]


# Canonical-LHSAA-xlsx-name → list of additional name strings to probe in
# the lhsaaonline data. Investigation 1 (2026-05-27) found Cohen College
# Prep listed under "Walter L. Cohen" in lhsaaonline reports. Engine
# data/school_aliases.py captures the same mapping for ingest; the
# pre-flight needs to apply it at lookup time so coverage attribution
# doesn't undercount due to renaming.
SAMPLE_LHSAA_NAME_ALIASES: dict[str, list[str]] = {
    "Cohen College Prep": ["Walter L. Cohen"],
}


def fetch_report(session: httpx.Client, sport: str, year: int, filter_value: str,
                 max_retries: int = 4) -> str:
    """Fetch one (sport, year, filter_value) report with retry-on-transient.

    Returns raw HTML. Retries DNS / connect / read errors with exponential
    backoff (1, 2, 4, 8 seconds). Raises the final exception if all retries
    fail."""
    cfg = SPORT_CFG[sport]
    data = {
        cfg["year_field"]: str(year),
        cfg["filter_field"]: filter_value,
        "resultdate": "", "n": "", "h": "", "f": "",
    }
    form_url = cfg["base_url"] + cfg["form_path"]
    report_url = cfg["base_url"] + "ReportSchedule.asp" + cfg["report_suffix"]
    last_err: Exception | None = None
    for attempt in range(max_retries):
        try:
            r = session.post(report_url, data=data,
                             headers={"Referer": form_url}, timeout=60)
            r.raise_for_status()
            return r.text
        except (httpx.ConnectError, httpx.ReadTimeout, httpx.ReadError,
                httpx.RemoteProtocolError) as e:
            last_err = e
            if attempt < max_retries - 1:
                backoff = 2 ** attempt  # 1, 2, 4, 8
                time.sleep(backoff)
                continue
            raise
    if last_err:
        raise last_err
    return ""


def count_school_games_in_report(html: str, school_name: str) -> int:
    """Count game rows (W/L) belonging to ``school_name`` in the report.

    Mirrors the existing scraper's table-walking logic: lhsaaonline renders
    one table per school with a header that contains 'School' + 'Win/Loss'.
    We find tables where the school column contains our target name and
    count rows with WL in {W, L}.
    """
    soup = BeautifulSoup(html, "html.parser")
    n_games = 0
    norm_target = school_name.strip().lower()
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if not rows:
            continue
        header = [td.get_text(strip=True) for td in rows[0].find_all(["td", "th"])]
        if "School" not in header or not any("Win/Loss" in h for h in header):
            continue
        # Find which column is the school
        try:
            school_col_idx = header.index("School")
        except ValueError:
            continue
        # WL column — varies; pick first that starts with Win/Loss
        wl_col_idx = next((i for i, h in enumerate(header) if "Win/Loss" in h), None)
        if wl_col_idx is None:
            continue
        for row in rows[1:]:
            cols = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
            if len(cols) < max(school_col_idx, wl_col_idx) + 1:
                continue
            school_in_row = cols[school_col_idx].strip().lower()
            if school_in_row != norm_target:
                continue
            wl = cols[wl_col_idx].strip().upper()
            if wl in ("W", "L"):
                n_games += 1
    return n_games


def main() -> int:
    print("=" * 70)
    print("Workstream B1.2b Pre-flight Coverage Check")
    print("=" * 70)

    # Load participation matrix from the committed xlsx
    print("\nLoading LHSAA Schools by Sport.xlsx...")
    wb = openpyxl.load_workbook(REPO_ROOT / "data/lhsaa/LHSAA Schools by Sport.xlsx",
                                  read_only=True, data_only=True)
    ws = wb["Schools by Sport"]

    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    sport_columns = {sp: header.index(sp) for sp in SPORT_CFG.keys() if sp in header}
    school_col = header.index("School")
    city_col = header.index("City")
    class_col = header.index("Classification")

    # Map (school_name) → {sport: True/False from 'X' flag} + class
    participation = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = (row[school_col] or "").strip()
        if not name:
            continue
        sports_fielded = {sp: bool(row[idx] == "X") for sp, idx in sport_columns.items()}
        participation[name] = {
            "city": row[city_col],
            "class": row[class_col],
            "sports": sports_fielded,
        }

    print(f"Loaded participation for {len(participation)} schools")
    print(f"Sport columns parsed: {list(sport_columns.keys())}")

    # ---------------------------------------------------------------------------
    # Union-sweep approach: for each (sport, year), fetch ALL filter_values
    # and build a (sport, year) → {school: games} dict. Then look up each
    # sample school against the dict. This avoids the brittle assumption
    # that a school's LHSAA classification matches the filter to use (Pickering
    # canonical 2A appears in GBB 5A report; Caddo Magnet canonical 4A appears
    # in Volleyball Division I — so per-(sport,year) sweep of all filters is
    # the robust approach).
    # ---------------------------------------------------------------------------
    print(f"\nSample schools: {[s[0] for s in SAMPLE_SCHOOLS]}")

    # Determine which sports + years to sweep
    sports_in_scope = set()
    for school_name, _, _ in SAMPLE_SCHOOLS:
        pdata = participation.get(school_name)
        if pdata is None:
            continue
        for sp, yes in pdata["sports"].items():
            if yes and sp != "Football":   # skip Football per Reese
                sports_in_scope.add(sp)

    # Filter values to sweep per filter_style
    CLASS_FILTERS = ["5A", "4A", "3A", "2A", "1A"]
    DIV_FILTERS = ["I", "II", "III", "IV", "V"]

    print(f"\nSweep plan: {len(sports_in_scope)} sports × {len(YEARS)} years × "
          f"5 filter_values per sport-year")
    print(f"Sports in scope: {sorted(sports_in_scope)}")

    # Build the union table: (sport, year) → {school_name: games_count}
    # Uses fixed column indices (school@1, wl@10) per the existing scraper's
    # _SCHEMA_12 / _SCHEMA_13 — header.index("School") returns the wrong
    # offset because the header row has extra section-header cells while
    # data rows have 12 (or 13) columns; both schemas put school @ 1 and
    # wl @ 10.
    union_table: dict[tuple[str, int], dict[str, int]] = {}
    fetch_count = 0
    SCHOOL_COL = 1
    WL_COL = 10

    # ---------------------------------------------------------------------------
    # Investigation 2 finding (2026-05-27): lhsaaonline silently returns the
    # PREVIOUSLY-FETCHED-GENDER's data when you reuse an httpx session across
    # Boys/Girls Basketball or Boys/Girls Soccer requests. The bb=1/bb=2 and
    # so=1/so=2 URL query params are ignored on cross-gender reuse. Fix:
    # open a FRESH httpx.Client PER SPORT. Mirrors the production scraper's
    # pattern (scripts/ingest_sports_historical.py:584 — run_sport opens its
    # own session). Existing production data is NOT corrupted because the
    # production scraper invokes one sport per run_sport call.
    # ---------------------------------------------------------------------------
    for sport in sorted(sports_in_scope):
        with httpx.Client(timeout=60, follow_redirects=True) as session:
            cfg = SPORT_CFG[sport]
            filters = DIV_FILTERS if cfg["filter_style"] == "division" else CLASS_FILTERS
            for year in YEARS:
                schools_for_sy: dict[str, int] = {}
                for fv in filters:
                    t_start = time.time()
                    try:
                        html = fetch_report(session, sport, year, fv)
                    except Exception as e:
                        print(f"  [{sport} {year} filter={fv}] FETCH ERR: {type(e).__name__}: {e}", flush=True)
                        time.sleep(REQUEST_DELAY)
                        continue
                    fetch_count += 1
                    elapsed = time.time() - t_start
                    print(f"  fetch {fetch_count:>3}/140  {sport:18} {year} filter={fv:<3}  "
                          f"{len(html)//1000}KB  {elapsed:.1f}s", flush=True)
                    soup = BeautifulSoup(html, "html.parser")
                    for table in soup.find_all("table"):
                        rows = table.find_all("tr")
                        if not rows: continue
                        header = [td.get_text(strip=True) for td in rows[0].find_all(["td", "th"])]
                        # Header presence check — same as scraper
                        if "School" not in header or not any("Win/Loss" in h for h in header):
                            continue
                        for row in rows[1:]:
                            cols = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
                            # Match _SCHEMA_12 / _SCHEMA_13: 12 or 13 cols, school @ 1, wl @ 10
                            if len(cols) not in (12, 13):
                                continue
                            sname = cols[SCHOOL_COL].strip()
                            wl = cols[WL_COL].strip().upper()
                            if sname and wl in ("W", "L"):
                                schools_for_sy[sname] = schools_for_sy.get(sname, 0) + 1
                    time.sleep(REQUEST_DELAY)
                union_table[(sport, year)] = schools_for_sy
                print(f"  [{sport} {year}] {len(schools_for_sy)} schools, "
                      f"{sum(schools_for_sy.values())} game-rows")

    print(f"\nTotal fetches: {fetch_count}  (cached across schools — same fetch covers all 5 samples)")

    # ---------------------------------------------------------------------------
    # Dump union_table to disk for Investigation 2 diagnosis
    # ---------------------------------------------------------------------------
    union_dump = {
        f"{sport}|{year}": {
            "n_schools": len(d),
            "n_games": sum(d.values()),
            "sample_schools": {k: d[k] for k in ["Mt. Carmel", "Cohen College Prep",
                                                  "Walter L. Cohen", "Pickering",
                                                  "Morris Jeff", "Caddo Magnet",
                                                  "Archbishop Chapelle"] if k in d},
        }
        for (sport, year), d in union_table.items()
    }
    dump_dir = REPO_ROOT / "reports" / "audits"
    dump_dir.mkdir(parents=True, exist_ok=True)
    (dump_dir / "workstream_b1_2b_preflight_union_table_dump.json").write_text(
        json.dumps(union_dump, indent=2, default=str)
    )
    print(f"Union table dump → reports/audits/workstream_b1_2b_preflight_union_table_dump.json")

    # ---------------------------------------------------------------------------
    # Look up each sample school against the union table
    # ---------------------------------------------------------------------------
    school_results = []
    for school_name, school_city, school_class in SAMPLE_SCHOOLS:
        print(f"\n--- {school_name} ({school_class}, {school_city}) ---")
        pdata = participation.get(school_name)
        if pdata is None:
            print(f"  NOT FOUND in LHSAA Schools by Sport.xlsx — skipping")
            school_results.append({
                "school": school_name, "class": school_class, "city": school_city,
                "sports_in_xlsx": None,
                "probes": [],
                "found_count": 0, "probe_count": 0, "coverage_pct": None,
                "skip_reason": "not in participation matrix",
            })
            continue

        sports_fielded = [sp for sp, yes in pdata["sports"].items() if yes]
        print(f"  Sports fielded per xlsx: {sports_fielded}")
        sports_to_probe = [sp for sp in sports_fielded if sp != "Football"]

        # Probe canonical name + any known lhsaaonline aliases
        probe_names = [school_name] + SAMPLE_LHSAA_NAME_ALIASES.get(school_name, [])
        probes = []
        for sport in sports_to_probe:
            for year in YEARS:
                schools_for_sy = union_table.get((sport, year), {})
                n_games = sum(schools_for_sy.get(n, 0) for n in probe_names)
                # Record which names hit
                hits = {n: schools_for_sy.get(n, 0) for n in probe_names if schools_for_sy.get(n, 0) > 0}
                probes.append({
                    "sport": sport, "year": year,
                    "games_found": n_games, "found": n_games > 0,
                    "matched_via": hits if hits else None,
                })

        n_probes = len(probes)
        n_found = sum(1 for p in probes if p["found"])
        coverage = n_found / n_probes if n_probes else 0.0
        print(f"  Probes: {n_probes}, sport-years with games found: {n_found}")
        print(f"  Coverage: {coverage*100:.1f}%")
        for p in probes:
            marker = "✓" if p["found"] else "·"
            print(f"    {marker} {p['sport']:20} {p['year']}  games={p['games_found']}")
        school_results.append({
            "school": school_name, "class": school_class, "city": school_city,
            "sports_in_xlsx": sports_fielded,
            "probes": probes,
            "found_count": n_found, "probe_count": n_probes,
            "coverage_pct": coverage,
        })

    # ---------------------------------------------------------------------------
    # Verdict
    # ---------------------------------------------------------------------------
    print("\n" + "=" * 70)
    print("Verdict")
    print("=" * 70)
    valid_results = [r for r in school_results if r["coverage_pct"] is not None]
    n_at_or_above_80 = sum(1 for r in valid_results if r["coverage_pct"] >= 0.80)
    n_total = len(valid_results)
    pass_threshold = n_total * 0.80
    passes = n_at_or_above_80 >= pass_threshold

    print(f"\nSchools sampled: {len(SAMPLE_SCHOOLS)}")
    print(f"Schools with usable participation data: {n_total}")
    print(f"Schools with ≥80% game-coverage in 2021-2024: {n_at_or_above_80}")
    print(f"Pass criterion: ≥{pass_threshold:.1f} schools at ≥80% coverage")
    print(f"\nTotal fetches made: {fetch_count}")
    print()
    for r in valid_results:
        cov = r["coverage_pct"] * 100 if r["coverage_pct"] else 0
        mark = "PASS" if cov >= 80 else "FAIL"
        print(f"  [{mark}] {r['school']:30}  class={r['class']:>3}  "
              f"sports={len(r['sports_in_xlsx'] or [])}  coverage={cov:.1f}%")

    print()
    if passes:
        print(f">>> VERDICT: PASS  ({n_at_or_above_80}/{n_total} schools at ≥80% coverage)")
        print(">>> Option (a) — full backfill 2021-2025 — LOCKS")
    else:
        print(f">>> VERDICT: FAIL  ({n_at_or_above_80}/{n_total} schools at ≥80% coverage)")
        print(">>> Fall back to Option (c) — hybrid scoping per-school")

    # ---------------------------------------------------------------------------
    # Artifacts
    # ---------------------------------------------------------------------------
    now = datetime.utcnow().isoformat() + "Z"
    findings = {
        "generated_utc": now,
        "samples": school_results,
        "n_total": n_total,
        "n_pass_at_80pct": n_at_or_above_80,
        "pass_threshold": pass_threshold,
        "verdict_passes": passes,
        "fetch_count": fetch_count,
    }
    out_dir = REPO_ROOT / "reports" / "audits"
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "workstream_b1_2b_preflight.json").write_text(
        json.dumps(findings, indent=2, default=str)
    )
    print(f"\nArtifact: reports/audits/workstream_b1_2b_preflight.json")
    return 0 if passes else 1


if __name__ == "__main__":
    sys.exit(main())

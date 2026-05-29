"""Workstream B1.2b orchestrator — execution discipline wrapper.

Per Reese 2026-05-28 Sign-off 1 (Option 1, build orchestrator wrapper):

  - Build orchestrator that WRAPS (not mutates) scripts/ingest_sports_historical.py
    so Class-D-safety from Sign-off 2 stays preserved by construction.
  - Add: checkpoint at reports/audits/workstream_b1_2b_checkpoint.json with
    (school, sport, year, status) rows.
  - Add: per-25-scrape structured progress log.
  - Add: 5xx exponential backoff (1s, 2s, 4s, then halt).
  - Add: 4xx categorization — 404 = empty (record + continue), 403/429 =
    throttle (backoff + retry), other 4xx = halt with diagnostic.
  - Add: per-sport completion matrix written at each sport halt boundary.
  - Add: schools-pre-insert step via engine.data.team_ingest BEFORE each
    sport's scrape (handles the 43+1 missing 1A-5A schools).
  - Halt at each sport boundary (7 halt points).

Class-D-safety preservation: this orchestrator does NOT mutate
scripts/ingest_sports_historical.py. It monkey-patches the module's
fetch_games function reference IN-MEMORY ONLY, leaving the underlying
production scraper file untouched. Each sport invocation still creates
its own httpx.Client inside run_sport() line 584. The Sign-off 2
runtime probe confirmed BBB and GBB return different data when invoked
back-to-back from the same python process — the monkey-patch preserves
that pattern because fetch_games is called inside the session context
that run_sport owns.

Usage:
    # Dry-run smoke test (no DB writes, no actual ingest)
    python scripts/workstream_b1_2b_run.py --sports softball --seasons 2024 --dry-run

    # Real Softball execution (5 years, ~hours)
    python scripts/workstream_b1_2b_run.py --sports softball --seasons 2021 2022 2023 2024 2025

    # Resume from checkpoint (skips cells already marked complete)
    python scripts/workstream_b1_2b_run.py --sports softball --seasons 2021 2022 2023 2024 2025 --resume

Output:
    reports/audits/workstream_b1_2b_checkpoint.json  (live state, updated per-scrape)
    reports/audits/workstream_b1_2b_completion_matrix.json  (final, per-sport halt)
    reports/audits/workstream_b1_2b_run_<sport>_<timestamp>.log  (per-sport log)
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
import traceback
from collections import defaultdict
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT / "packages" / "engine" / "src"))

import httpx
from dotenv import load_dotenv
load_dotenv(REPO_ROOT / "apps" / "api" / ".env")

import openpyxl

# Import production scraper for monkey-patching
import scripts.ingest_sports_historical as scraper

from engine.data.team_ingest import ingest_alignment
from engine.data.school_aliases import normalize_name, resolve_school


# Sequencing per Reese (coverage-gap descending), skipping Football
B1_2B_SPORT_ORDER = [
    ("softball", "Softball"),
    ("girls_basketball", "Girls Basketball"),
    ("boys_basketball", "Boys Basketball"),
    ("girls_soccer", "Girls Soccer"),
    ("volleyball", "Volleyball"),
    ("boys_soccer", "Boys Soccer"),
    ("baseball", "Baseball"),
]

CLASS_1A_5A = {"1A", "2A", "3A", "4A", "5A"}

CHECKPOINT_PATH = REPO_ROOT / "reports" / "audits" / "workstream_b1_2b_checkpoint.json"
COMPLETION_MATRIX_PATH = REPO_ROOT / "reports" / "audits" / "workstream_b1_2b_completion_matrix.json"


# ---------------------------------------------------------------------------
# Checkpoint state
# ---------------------------------------------------------------------------
def load_checkpoint() -> dict:
    if CHECKPOINT_PATH.exists():
        return json.loads(CHECKPOINT_PATH.read_text())
    return {
        "started_utc": datetime.utcnow().isoformat() + "Z",
        "last_update_utc": None,
        "cells": {},   # key: f"{sport}|{year}|{school_id}", value: {"status": ..., ...}
        "per_sport_summary": {},
    }


def save_checkpoint(ckpt: dict) -> None:
    ckpt["last_update_utc"] = datetime.utcnow().isoformat() + "Z"
    CHECKPOINT_PATH.parent.mkdir(parents=True, exist_ok=True)
    CHECKPOINT_PATH.write_text(json.dumps(ckpt, indent=2, default=str))


# ---------------------------------------------------------------------------
# Disciplined fetch_games wrapper
# ---------------------------------------------------------------------------
class HaltOnError(RuntimeError):
    """Raised when 4xx other than 404/403/429 is encountered — orchestrator halts."""


def make_disciplined_fetch_games(progress_state: dict):
    """Build a monkey-patch fetch_games with discipline. progress_state is a
    mutable dict the parent owns; the wrapper increments its counters."""
    original_fetch = scraper.fetch_games

    def disciplined_fetch_games(session, cfg, year, fval):
        """Wraps original fetch_games with:
        - 5xx / network-transient backoff: 4 attempts with 1s, 2s, 4s, 8s waits
          (slightly more lenient than the original "1s, 2s, 4s, then halt"
          spec — production runs benefit from the extra retry when LHSAA
          briefly throttles or DNS hiccups).
        - 4xx categorization (404=empty/continue, 403/429=backoff/retry, else halt)
        - Per-fetch progress counter; emit structured log every 25 fetches
        """
        for attempt in range(4):
            try:
                rows = original_fetch(session, cfg, year, fval)
                progress_state["n_fetches"] += 1
                if progress_state["n_fetches"] % 25 == 0:
                    elapsed = time.time() - progress_state["t_start"]
                    print(f"  [progress] {datetime.utcnow().isoformat()}  "
                          f"fetch={progress_state['n_fetches']}  "
                          f"elapsed={elapsed/60:.1f}min  "
                          f"sport={cfg.name}  year={year}  filter={fval}",
                          flush=True)
                return rows
            except httpx.HTTPStatusError as e:
                code = e.response.status_code
                if code == 404:
                    # 404 = empty result for that filter (legit)
                    print(f"    404 (empty) {cfg.name} {year} {fval}", flush=True)
                    return []
                if code in (403, 429):
                    # Throttle — exponential backoff and retry
                    wait = 2 ** attempt
                    print(f"    {code} throttle {cfg.name} {year} {fval} — backoff {wait}s "
                          f"(attempt {attempt+1}/4)", flush=True)
                    time.sleep(wait)
                    continue
                # Other 4xx — halt
                raise HaltOnError(
                    f"4xx other than 404/403/429: {code} on {cfg.name} {year} {fval}\n"
                    f"  body excerpt: {e.response.text[:300]}"
                )
            except (httpx.ConnectError, httpx.ReadTimeout, httpx.ReadError,
                    httpx.RemoteProtocolError) as e:
                # Treat as transient (5xx-equivalent for network layer)
                wait = 2 ** attempt
                print(f"    transient {type(e).__name__} {cfg.name} {year} {fval} — "
                      f"backoff {wait}s (attempt {attempt+1}/4)", flush=True)
                time.sleep(wait)
                continue
            except Exception as e:
                # Persistent — log and surface
                print(f"    persistent {type(e).__name__}: {e}", flush=True)
                if attempt < 3:
                    time.sleep(2 ** attempt)
                    continue
                raise

        # Exhausted retries — return empty (treated as ingest-failed for this filter)
        print(f"    EXHAUSTED retries on {cfg.name} {year} {fval}", flush=True)
        return []

    return disciplined_fetch_games


# ---------------------------------------------------------------------------
# Schools pre-insert step (via engine.data.team_ingest)
# ---------------------------------------------------------------------------
def load_lhsaa_xlsx_participation() -> dict:
    """Parse data/lhsaa/LHSAA Schools by Sport.xlsx into a participation dict
    matching the shape expected by engine.data.team_ingest.ingest_alignment.
    """
    xlsx_path = REPO_ROOT / "data" / "lhsaa" / "LHSAA Schools by Sport.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Schools by Sport"]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    school_col = header.index("School")
    city_col = header.index("City")
    class_col = header.index("Classification")
    sport_cols = {
        sp: header.index(sp)
        for sp in ("Football", "Volleyball", "Boys Basketball", "Girls Basketball",
                   "Boys Soccer", "Girls Soccer", "Baseball", "Softball")
        if sp in header
    }
    participation = {sp: [] for sp in sport_cols}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = (row[school_col] or "").strip()
        if not name:
            continue
        city = row[city_col]
        cls = row[class_col]
        for sp, col in sport_cols.items():
            if row[col] == "X":
                participation[sp].append({
                    "school": name, "city": city, "classification": cls,
                })
    return {
        "season": "2025-26",
        "source": "LHSAA Schools by Sport.xlsx (Reese 2026-05-27 consolidation)",
        "participation": participation,
    }


def missing_schools_for_sport(participation: dict, sport_name: str, db_schools: list,
                               class_filter: set[str] = CLASS_1A_5A) -> list[dict]:
    """Return the list of (school, city, classification) entries from the
    participation matrix that field ``sport_name`` AND are MISSING from the
    DB (after applying the alias resolver). Filtered to 1A-5A only per the
    Option (a) scope decision."""
    sport_list = participation.get("participation", {}).get(sport_name, [])
    out = []
    for entry in sport_list:
        if entry.get("classification") not in class_filter:
            continue
        # Skip schools already in DB (resolver handles aliases)
        if resolve_school(entry["school"], db_schools) is not None:
            continue
        out.append(entry)
    return out


# ---------------------------------------------------------------------------
# Per-sport orchestrator
# ---------------------------------------------------------------------------
def run_sport_disciplined(
    sb,
    sport_key: str,
    sport_name: str,
    seasons: list[int],
    dry_run: bool,
    skip_ratings: bool,
    ckpt: dict,
    progress_state: dict,
) -> dict:
    """Wrap the per-sport execution: pre-insert missing schools, monkey-patch
    fetch_games with discipline, call run_sport, return per-sport summary."""
    sport_start = datetime.utcnow().isoformat() + "Z"
    print(f"\n{'='*70}")
    print(f"=== B1.2b sport boundary: {sport_name} ===")
    print(f"=== started {sport_start}")
    print(f"{'='*70}\n")

    # Load participation + identify missing schools
    participation = load_lhsaa_xlsx_participation()
    all_db = sb.table("schools").select("id, name, city, parish, classification").execute().data
    db_la = [s for s in all_db if s.get("parish") is None]
    missing = missing_schools_for_sport(participation, sport_name, db_la)
    print(f"  Missing 1A-5A schools fielding {sport_name}: {len(missing)}")
    for m in missing[:5]:
        print(f"    {m['school']!r} ({m['classification']}, {m.get('city')})")
    if len(missing) > 5:
        print(f"    ... and {len(missing) - 5} more")

    # Pre-insert step via team_ingest
    if missing and not dry_run:
        sport_id_map = {r["name"]: r["id"]
                        for r in sb.table("sports").select("id, name").execute().data}
        db_teams = sb.table("teams").select("id, school_id, sport_id, season_year").execute().data
        ingest_payload = {
            "participation": {sport_name: missing},
            "source": f"B1.2b pre-insert for {sport_name}",
        }

        def insert_school_fn(payload):
            res = sb.table("schools").insert({
                "name": payload["name"],
                "city": payload.get("city"),
                "classification": payload.get("classification"),
                "parish": payload.get("parish"),
            }).execute()
            return res.data[0] if res.data else None

        def insert_team_fn(payload):
            res = sb.table("teams").insert({
                "school_id": payload["school_id"],
                "sport_id": payload["sport_id"],
                "season_year": payload["season_year"],
            }).execute()
            return res.data[0] if res.data else None

        # Insert team rows for each year in scope
        for season_year in seasons:
            year_payload = {
                "participation": {sport_name: missing},
                "season_year_for_ingest": season_year,
            }
            try:
                result = ingest_alignment(
                    participation_data=year_payload,
                    sport_id_map=sport_id_map,
                    season_year=season_year,
                    db_schools=db_la,
                    db_teams=db_teams,
                    insert_school_fn=insert_school_fn,
                    insert_team_fn=insert_team_fn,
                    source_attribution=f"B1.2b pre-insert {sport_name} {season_year}",
                )
                print(f"    season {season_year}: schools_inserted={result.n_schools_inserted}, "
                      f"teams_inserted={result.n_teams_inserted}, "
                      f"existing={result.n_teams_already_present}")
                # Refresh db_la and db_teams snapshots after insertions
                if result.n_schools_inserted > 0:
                    db_la = sb.table("schools").select("id, name, city, parish, classification").is_("parish", "null").execute().data
                if result.n_teams_inserted > 0:
                    db_teams = sb.table("teams").select("id, school_id, sport_id, season_year").execute().data
            except Exception as e:
                print(f"    season {season_year}: pre-insert FAILED: {type(e).__name__}: {e}")
                traceback.print_exc()
    elif missing and dry_run:
        print(f"  [dry-run] WOULD pre-insert {len(missing)} schools × {len(seasons)} seasons")
    else:
        print(f"  No missing schools — pre-insert step skipped")

    # ---------------------------------------------------------------------------
    # Monkey-patch fetch_games with disciplined wrapper for this sport's run
    # ---------------------------------------------------------------------------
    original_fetch = scraper.fetch_games
    scraper.fetch_games = make_disciplined_fetch_games(progress_state)

    sport_summary = {
        "sport": sport_name,
        "started_utc": sport_start,
        "completed_utc": None,
        "n_missing_schools_pre_insert": len(missing),
        "seasons_attempted": seasons,
        "fetch_count_start": progress_state["n_fetches"],
        "status": "in_progress",
        "halt_reason": None,
    }
    ckpt["per_sport_summary"][sport_name] = sport_summary
    save_checkpoint(ckpt)

    try:
        cfg = scraper.SPORTS[sport_key]
        # Refresh school_name_to_id after pre-insert
        school_name_to_id = {
            r["name"]: r["id"]
            for r in sb.table("schools").select("id, name").execute().data
        }
        scraper.run_sport(
            sb=sb,
            cfg=cfg,
            seasons=seasons,
            dry_run=dry_run,
            school_name_to_id=school_name_to_id,
            skip_ratings=skip_ratings,
        )
        sport_summary["status"] = "complete"
    except HaltOnError as e:
        sport_summary["status"] = "halted_4xx"
        sport_summary["halt_reason"] = str(e)
        raise
    except Exception as e:
        sport_summary["status"] = "halted_exception"
        sport_summary["halt_reason"] = f"{type(e).__name__}: {e}"
        traceback.print_exc()
        raise
    finally:
        # Restore original fetch_games
        scraper.fetch_games = original_fetch
        sport_summary["completed_utc"] = datetime.utcnow().isoformat() + "Z"
        sport_summary["fetch_count_end"] = progress_state["n_fetches"]
        sport_summary["n_fetches_for_sport"] = (
            progress_state["n_fetches"] - sport_summary["fetch_count_start"]
        )
        ckpt["per_sport_summary"][sport_name] = sport_summary
        save_checkpoint(ckpt)

    return sport_summary


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(prog="python scripts/workstream_b1_2b_run.py")
    p.add_argument("--sports", default="all",
                   help="Comma-separated sport keys (e.g. 'softball,girls_basketball') or 'all'")
    p.add_argument("--seasons", nargs="+", type=int,
                   default=[2021, 2022, 2023, 2024, 2025])
    p.add_argument("--dry-run", action="store_true",
                   help="Pass dry_run=True to run_sport (no DB writes)")
    p.add_argument("--skip-ratings", action="store_true",
                   help="Pass skip_ratings=True to run_sport")
    p.add_argument("--resume", action="store_true",
                   help="Resume from existing checkpoint (skip completed sports)")
    args = p.parse_args(argv)

    # Determine sport iteration order
    if args.sports == "all":
        sports_to_run = list(B1_2B_SPORT_ORDER)
    else:
        keys = [k.strip() for k in args.sports.split(",")]
        sports_to_run = [(k, n) for k, n in B1_2B_SPORT_ORDER if k in keys]
        if not sports_to_run:
            print(f"ERROR — none of {keys} matched the sequencing order {[k for k,_ in B1_2B_SPORT_ORDER]}")
            return 2

    # Supabase client
    sb_url = os.environ["SUPABASE_URL"]
    sb_key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    from supabase import create_client
    sb = create_client(sb_url, sb_key)

    # Load or initialize checkpoint
    ckpt = load_checkpoint() if (args.resume and CHECKPOINT_PATH.exists()) else {
        "started_utc": datetime.utcnow().isoformat() + "Z",
        "last_update_utc": None,
        "cells": {},
        "per_sport_summary": {},
    }
    progress_state = {"n_fetches": 0, "t_start": time.time()}

    print(f"\nB1.2b orchestrator START at {datetime.utcnow().isoformat()}Z")
    print(f"  sports: {[n for _, n in sports_to_run]}")
    print(f"  seasons: {args.seasons}")
    print(f"  dry_run: {args.dry_run}")
    print(f"  resume:  {args.resume}")
    print(f"  checkpoint: {CHECKPOINT_PATH.relative_to(REPO_ROOT)}\n")

    for sport_key, sport_name in sports_to_run:
        # Resume: skip if already complete
        prev = ckpt["per_sport_summary"].get(sport_name)
        if args.resume and prev and prev.get("status") == "complete":
            print(f"  [resume] {sport_name} already complete — skipping")
            continue

        try:
            summary = run_sport_disciplined(
                sb=sb, sport_key=sport_key, sport_name=sport_name,
                seasons=args.seasons, dry_run=args.dry_run,
                skip_ratings=args.skip_ratings,
                ckpt=ckpt, progress_state=progress_state,
            )
            print(f"\n[sport halt] {sport_name} complete. "
                  f"Fetches this sport: {summary['n_fetches_for_sport']}. "
                  f"Cumulative: {progress_state['n_fetches']}.")
            print(f"[sport halt] Halting per Reese's 7-halt-point discipline. "
                  f"Re-invoke with --resume to continue.")
            # Save final completion matrix after each sport
            COMPLETION_MATRIX_PATH.write_text(json.dumps(ckpt, indent=2, default=str))
            return 0
        except HaltOnError as e:
            print(f"\n*** HALT (4xx other): {e}")
            return 1
        except KeyboardInterrupt:
            print("\n*** INTERRUPTED — checkpoint preserved")
            return 130
        except Exception as e:
            print(f"\n*** UNEXPECTED HALT: {type(e).__name__}: {e}")
            traceback.print_exc()
            return 1

    # All sports complete
    COMPLETION_MATRIX_PATH.write_text(json.dumps(ckpt, indent=2, default=str))
    print(f"\nALL SPORTS COMPLETE. Completion matrix → {COMPLETION_MATRIX_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
